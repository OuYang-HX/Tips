import pyautogui
import time
import pyperclip
import openpyxl

def get_code_description(query_contents):
    # # 设置延迟以确保程序有足够的时间执行操作
    # pyautogui.PAUSE = 1

    # # 移动鼠标到地址栏的位置并点击
    # pyautogui.click(300, 80)

    # # 输入网址 https://chat.openai.com/
    # pyautogui.write("https://chat.openai.com/")
    # time.sleep(1)
    # pyautogui.press("enter", 2)

    # # 等待页面加载
    # time.sleep(5)

    # 移动鼠标到搜索框的位置并点击
    pyautogui.click(1015, 1325)

    # 查询
    # 粘贴剪贴板内容到搜索框
    pyperclip.copy("")  # 清空剪贴板，避免在粘贴时同时触发搜索
    # 将 query_contents 复制到剪贴板
    pyperclip.copy(query_contents)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press("enter")

    # 等待查询结果加载
    time.sleep(10)

    # 指定要搜索的图片路径
    image_path = "D:\大模型\白盒案例收集\复制图标.png"

    
    try:
        # 查找指定图片在屏幕上的位置
        image_location = pyautogui.locateOnScreen(image_path)
        if image_location is not None:
            # 图片存在于屏幕上，获取图片的中心点坐标
            image_center = pyautogui.center(image_location)
        
            # 点击图片的中心点坐标
            pyautogui.click(image_center)

            # 获取剪贴板的内容
            clipboard_content = pyperclip.paste()
            print("剪贴板内容:", clipboard_content)
            return clipboard_content
        else:
            print("未找到指定图片")
            return "ERROR"
    except pyautogui.ImageNotFoundException:
        # 图片未找到
        print("图片未找到")
        return "ERROR"

# 打开 Excel 文件
workbook = openpyxl.load_workbook("D:\大模型\白盒案例收集\错误代码示例说明测试.xlsx")

# 选择要读取的工作表
sheet = workbook.active

# 如果第三列不存在，则添加第三列
if sheet.max_column < 3:
    sheet.insert_cols(3)

# 遍历每一行，并读取每一列的内容
for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    if row_number > 1:  # 跳过表头
        question_type = row[0].strip()  # 问题类型
        code_example = row[1].strip()   # 代码示例
        query_contents = "用一段话（200字以内）解释并总结下面代码存在" + question_type + "风险的原因:\n" + code_example
        
        code_description = get_code_description(query_contents)
        
        # 将代码示例说明写入第三列
        sheet.cell(row=row_number, column=3, value=code_description)

# 保存修改后的 Excel 文件
workbook.save("D:\大模型\白盒案例收集\错误代码示例说明测试.xlsx")

# 关闭 Excel 文件
workbook.close()
