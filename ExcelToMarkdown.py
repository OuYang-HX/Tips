from openpyxl import load_workbook
import argparse

def is_code_block(cell_value):
    return cell_value.strip().endswith("}") or "```" in cell_value

def generate_markdown_from_excel(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    markdown_content = ""

    # 获取列名
    column_names = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
        first_cell_value = str(row[0])
        markdown_content += f"# {first_cell_value}\n"
        for i in range(1, len(row)):
            if row[i] is not None:
                cell_value = str(row[i])
                if column_names[i] == "正确代码示例" or column_names[i] == "错误代码示例":
                    markdown_content += f"## {column_names[i]}\n```java\n{cell_value}\n```\n"
                else:
                    if is_code_block(cell_value):
                        markdown_content += f"## {column_names[i]}\n```java\n{cell_value}\n```\n"
                    else:
                        markdown_content += f"## {column_names[i]}\n{cell_value}\n"

    return markdown_content


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Convert Excel to Markdown')
    parser.add_argument('--input', type=str, help='Input Excel file path', default='数据集.xlsx')
    parser.add_argument('--output', type=str, help='Output Markdown file path', default='数据集_reverse.md')
    args = parser.parse_args()

    excel_file = args.input
    markdown_output_file = args.output

    markdown_content = generate_markdown_from_excel(excel_file)

    with open(markdown_output_file, 'w', encoding='utf-8') as file:
        file.write(markdown_content)

    print(f"Markdown file has been generated and saved to {markdown_output_file}")
