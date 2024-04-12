# 导入所需模块
from openpyxl import Workbook  # 用于创建Excel工作簿
from openpyxl.utils import get_column_letter  # 用于获取Excel列字母表示
import argparse  # 用于解析命令行参数

# 读取Markdown文件内容并返回行列表
def read_markdown_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # 去除每行末尾的换行符
    lines = [line.rstrip() for line in lines]

    return lines

# 将Markdown内容解析成Excel格式
def parse_markdown_to_excel(markdown_lines):
    # 创建一个Excel工作簿和工作表
    wb = Workbook()
    ws = wb.active

    current_row = 1  # 当前行数
    current_column = 1  # 当前列数
    current_title = ''  # 当前标题
    titles = {}  # 标题字典，用于记录标题所在的行数

    in_code_block = False  # 标志是否处于代码块中
    code_block_lines = []  # 用于存储代码块内容的列表

    # 单元格内容字典，用于存储每个单元格的内容
    cell_contents = {}

    # 遍历Markdown内容的每一行
    for line in markdown_lines:
        if line.startswith('# '):  # 如果是一级标题
            current_title = line.strip('# ').strip()  # 获取标题内容
            current_row += 1  # 行数加1
            current_column = 1  # 列数重置为1
            ws.cell(row=current_row, column=current_column, value=current_title)  # 在当前位置写入标题内容
            titles[current_title] = current_row  # 记录标题所在的行数
        elif line.startswith('## '):  # 如果是二级标题
            sub_title = line.strip('## ').strip()  # 获取二级标题内容
            current_column += 1  # 列数加1
            ws.cell(row=1, column=current_column, value=sub_title)  # 在当前位置写入二级标题内容
        elif line.startswith('```'):  # 如果是代码块开始标记
            if in_code_block:  # 如果已经处于代码块中
                in_code_block = False  # 标志重置为False
                code_content = '\n'.join(code_block_lines)  # 将代码块内容列表连接成字符串
                if current_title in titles:  # 如果当前标题已存在于字典中
                    current_row = titles[current_title]  # 获取标题所在的行数
                    cell_coordinate = (current_row, current_column)  # 获取单元格坐标
                    if cell_coordinate in cell_contents:  # 如果单元格内容字典中已存在该单元格的内容
                        cell_contents[cell_coordinate] += "\n\n" + code_content  # 在原内容后面添加代码块内容
                    else:
                        cell_contents[cell_coordinate] = code_content  # 直接写入代码块内容
                code_block_lines = []  # 重置代码块内容列表
            else:
                in_code_block = True  # 标志设为True，表示进入代码块
        else:
            if in_code_block:  # 如果处于代码块中
                code_block_lines.append(line)  # 将行内容添加到代码块内容列表中
            else:
                if current_title in titles:  # 如果当前标题已存在于字典中
                    current_row = titles[current_title]  # 获取标题所在的行数
                    cell_coordinate = (current_row, current_column)  # 获取单元格坐标
                    if cell_coordinate in cell_contents:  # 如果单元格内容字典中已存在该单元格的内容
                        cell_contents[cell_coordinate] += "\n" + line  # 在原内容后面添加当前行内容
                    else:
                        cell_contents[cell_coordinate] = line  # 直接写入当前行内容
    
    # 根据单元格内容字典写入Excel单元格
    for coordinate, content in cell_contents.items():
        row, column = coordinate
        content = content.strip()  # 删除内容前后的空白字符
        if content:  # 检查内容是否为空
            ws.cell(row=row, column=column, value=content)
    
    # 设置单元格宽度
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)  # 获取列字母表示
        ws.column_dimensions[col_letter].width = 22  # 设置列宽度为22

    return wb  # 返回生成的Excel工作簿对象

# 主程序入口
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Convert Markdown to Excel')  # 创建解析器对象
    parser.add_argument('--input', type=str, help='Input Markdown file path', default='数据集.md')  # 添加命令行参数--input，用于指定输入Markdown文件路径，默认为“数据集.md”
    parser.add_argument('--output', type=str, help='Output Excel file path', default='数据集.xlsx')  # 添加命令行参数--output，用于指定输出Excel文件路径，默认为“数据集.xlsx”
    args = parser.parse_args()  # 解析命令行参数

    markdown_file = args.input  # 获取输入Markdown文件路径
    excel_file = args.output  # 获取输出Excel文件路径

    markdown_lines = read_markdown_file(markdown_file)  # 读取Markdown文件内容
    wb = parse_markdown_to_excel(markdown_lines)  # 将Markdown内容解析成Excel格式
    wb.save(excel_file)  # 保存生成的Excel文件
    print(f"Excel file has been generated and saved to {excel_file}")  # 打印提示信息，显示Excel文件已生成并保存的路径
